"""
excel_writer.py — Creates and maintains the job_tracker.xlsx file.

Sheet layout:
  Sheet 1 — "All Jobs"   : every qualifying listing ever found, chronological
  Sheet 2 — "Dashboard"  : auto-computed summary stats via Excel formulas

Columns (All Jobs sheet):
  A  Date Found     | B  Date Posted  | C  Title
  D  Company        | E  Source       | F  Work Type
  G  Location       | H  Match Score  | I  Priority
  J  Key Skills     | K  URL          | L  Status
  M  Query          | N  Notes

Color legend (row background):
  Green  (#C6EFCE) — Priority: High   (score ≥ 8)
  Yellow (#FFEB9C) — Priority: Medium (score 5-7)
  White             — Priority: Low   (score 3-4)
  Grey   (#D9D9D9) — Status: Applied / Closed
"""

import os
import logging
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_DATE_DATETIME

from config import EXCEL_FILE, MIN_SCORE
from scorer import priority_label

logger = logging.getLogger(__name__)

# ── Colour palette ─────────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1F3864"   # dark navy
CLR_HEADER_FG   = "FFFFFF"
CLR_HIGH_BG     = "C6EFCE"   # green
CLR_MED_BG      = "FFEB9C"   # yellow
CLR_LOW_BG      = "FFFFFF"   # white
CLR_ALT_ROW     = "F2F2F2"   # light grey alternate rows
CLR_APPLIED_BG  = "D9D9D9"   # grey — already applied

COL_HEADERS = [
    "Date Found", "Date Posted", "Title", "Company",
    "Source", "Work Type", "Location", "Match Score",
    "Priority", "Key Skills", "URL", "Status",
    "Query", "Notes",
]

COL_WIDTHS = [13, 13, 38, 28, 16, 16, 22, 13, 10, 40, 55, 14, 28, 30]

# ── Style helpers ──────────────────────────────────────────────────────────────

def _header_style() -> tuple[Font, PatternFill, Alignment]:
    font  = Font(name="Arial", bold=True, color=CLR_HEADER_FG, size=10)
    fill  = PatternFill("solid", fgColor=CLR_HEADER_BG)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, align


def _thin_border() -> Border:
    thin = Side(border_style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _row_fill(priority: str) -> PatternFill | None:
    mapping = {
        "High":   PatternFill("solid", fgColor=CLR_HIGH_BG),
        "Medium": PatternFill("solid", fgColor=CLR_MED_BG),
    }
    return mapping.get(priority)


# ── Workbook initialisation ────────────────────────────────────────────────────

def _create_workbook() -> Workbook:
    """Create a brand-new tracker workbook with headers and dashboard."""
    wb = Workbook()

    # ── All Jobs sheet ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Jobs"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True
    ws.row_dimensions[1].height = 30

    hfont, hfill, halign = _header_style()
    border = _thin_border()

    for col_idx, (header, width) in enumerate(zip(COL_HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font    = hfont
        cell.fill    = hfill
        cell.alignment = halign
        cell.border  = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Dashboard sheet ───────────────────────────────────────────────────────
    dash = wb.create_sheet("Dashboard")
    dash["A1"] = "Job Tracker — Dashboard"
    dash["A1"].font = Font(name="Arial", bold=True, size=14, color=CLR_HEADER_BG)

    stats = [
        ("Total listings found:",  "='All Jobs'!A1"),   # placeholder; updated below
        ("High Priority:",         "=COUNTIF('All Jobs'!I:I,\"High\")"),
        ("Medium Priority:",       "=COUNTIF('All Jobs'!I:I,\"Medium\")"),
        ("Low Priority:",          "=COUNTIF('All Jobs'!I:I,\"Low\")"),
        ("Applied:",               "=COUNTIF('All Jobs'!L:L,\"Applied\")"),
        ("Sources — LinkedIn:",    "=COUNTIF('All Jobs'!E:E,\"LinkedIn\")"),
        ("Sources — Wuzzuf:",      "=COUNTIF('All Jobs'!E:E,\"Wuzzuf\")"),
        ("Sources — Bayt:",        "=COUNTIF('All Jobs'!E:E,\"Bayt\")"),
        ("Sources — RemoteOK:",    "=COUNTIF('All Jobs'!E:E,\"RemoteOK\")"),
        ("Sources — WWR:",         "=COUNTIF('All Jobs'!E:E,\"WeWorkRemotely\")"),
        ("Sources — Himalayas:",   "=COUNTIF('All Jobs'!E:E,\"Himalayas\")"),
        ("Sources — Jobicy:",      "=COUNTIF('All Jobs'!E:E,\"Jobicy\")"),
    ]

    real_count_formula = "=COUNTA('All Jobs'!A:A)-1"  # subtract header row
    stats[0] = ("Total listings found:", real_count_formula)

    for row_offset, (label, formula) in enumerate(stats, start=3):
        dash.cell(row=row_offset, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        cell = dash.cell(row=row_offset, column=2, value=formula)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center")

    dash.column_dimensions["A"].width = 28
    dash.column_dimensions["B"].width = 16

    return wb


# ── Public interface ────────────────────────────────────────────────────────────

def append_jobs(jobs: list[dict]) -> int:
    """
    Append a list of scored job dicts to the Excel tracker.

    Returns the number of rows actually written (i.e. not duplicates, score ≥ MIN_SCORE).
    Job dicts must contain keys produced by scorer.score_job() / extract_key_requirements().
    """
    if not jobs:
        return 0

    # Load or create workbook
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb["All Jobs"]
    else:
        logger.info("Creating new workbook at %s", EXCEL_FILE)
        wb = _create_workbook()
        ws = wb["All Jobs"]

    # Collect existing URLs to skip duplicates within this write call
    existing_urls: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[10]:  # column K = URL (0-indexed: index 10)
            existing_urls.add(str(row[10]).strip())

    today = date.today().isoformat()
    border = _thin_border()
    written = 0

    for job in jobs:
        score    = job.get("score", 0)
        priority = priority_label(score)
        url      = job.get("url", "").strip()

        # Skip low-score or already-seen entries
        if score < MIN_SCORE:
            continue
        if url and url in existing_urls:
            continue

        # Determine next empty row
        next_row = ws.max_row + 1

        row_values = [
            today,
            job.get("date_posted", ""),
            job.get("title", ""),
            job.get("company", ""),
            job.get("source", ""),
            job.get("work_type", ""),
            job.get("location", ""),
            score,
            priority,
            job.get("key_requirements", ""),
            url,
            "Not Applied",
            job.get("query", ""),
            "",  # Notes — blank, filled by user
        ]

        fill = _row_fill(priority)

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in {3, 10, 11, 14}))
            cell.border    = border
            if fill:
                cell.fill = fill
            # Make URL clickable
            if col_idx == 11 and url:
                cell.hyperlink = url
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

        ws.row_dimensions[next_row].height = 28
        existing_urls.add(url)
        written += 1

    wb.save(EXCEL_FILE)
    logger.info("Excel | %d new rows written → %s", written, EXCEL_FILE)
    return written
